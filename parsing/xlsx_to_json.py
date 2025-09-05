#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import argparse
from pathlib import Path
from openpyxl import load_workbook


def pick(d: dict, keys: list[str], default=None):
    for k in keys:
        if k in d and d[k] not in (None, ""):
            return d[k]
    return default


def coerce_int(value):
    try:
        return int(str(value).strip())
    except Exception:
        return None


def row_to_property(row: dict, idx: int) -> dict:
    id_candidates = [
        "id", "property_id", "json_project_number", "matched_project_id",
        "project_number", "number",
    ]
    title_candidates = [
        "title", "project_name_en", "project_name_en_out", "json_project_name",
        "matched_project_name", "project_name", "project_name_en_tr", "project_name_ar",
    ]
    address_candidates = [
        "displayAddress", "display_address", "address", "area", "location",
        "area_name_en", "area_name_en_out", "area_name_en_tr",
    ]
    url_candidates = ["url", "share_url", "link", "pf_url"]
    price_candidates = ["actual_worth", "rent_value", "price", "price_value", "amount"]
    currency_candidates = ["priceCurrency", "currency"]
    bedrooms_candidates = ["bedrooms", "rooms", "bedroom", "rooms_en"]
    bathrooms_candidates = ["bathrooms", "bathroom"]
    size_candidates = ["sizeMin", "size", "area_sqm", "area_sqft", "procedure_area"]

    pid = pick(row, id_candidates)
    if pid is None:
        pid = f"XLSX_{idx+1}"

    title = pick(row, title_candidates, default="")
    address = pick(row, address_candidates, default="")
    url = pick(row, url_candidates, default="")
    price = pick(row, price_candidates)
    currency = pick(row, currency_candidates, default="AED")
    bedrooms = pick(row, bedrooms_candidates)
    bathrooms = pick(row, bathrooms_candidates)
    size_val = pick(row, size_candidates)

    trans_group = (str(row.get("trans_group_en") or row.get("procedure_name_en") or "")).lower()
    price_duration = "rent" if "rent" in trans_group else "sell"

    obj = {
        "id": pid,
        "url": url,
        "title": title,
        "displayAddress": address,
        "price": price,
        "priceCurrency": currency,
        "priceDuration": price_duration,
    }

    if bedrooms is not None:
        obj["bedrooms"] = coerce_int(bedrooms)
    if bathrooms is not None:
        obj["bathrooms"] = coerce_int(bathrooms)
    if size_val not in (None, ""):
        try:
            f = float(str(size_val).replace(",", "."))
            obj["sizeMin"] = f"{f} sqm"
        except Exception:
            obj["sizeMin"] = str(size_val)

    # Try parse bedrooms from text like "2 B/R"
    if obj.get("bedrooms") is None and isinstance(bedrooms, str):
        import re as _re
        m = _re.search(r"(\d+)", bedrooms)
        if m:
            obj["bedrooms"] = coerce_int(m.group(1))

    return obj


def read_xlsx(path: Path) -> list[dict]:
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = []
    for r in rows[1:]:
        row = {headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))}
        data.append(row)
    return data


def convert(xlsx_path: Path, out_path: Path):
    raw = read_xlsx(xlsx_path)
    items = [row_to_property(row, i) for i, row in enumerate(raw)]
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(items)} items to {out_path}")


def main():
    p = argparse.ArgumentParser(description="Convert XLSX to JSON for import_properties")
    p.add_argument("input_xlsx", type=Path)
    p.add_argument("output_json", type=Path)
    args = p.parse_args()
    convert(args.input_xlsx, args.output_json)


if __name__ == "__main__":
    main()


